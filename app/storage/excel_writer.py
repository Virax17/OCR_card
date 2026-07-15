from __future__ import annotations

from io import BytesIO
from typing import Callable

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PillowImage

from app.config import EXCEL_COLUMNS, EXCEL_HEADERS
from app.extraction.field_resolver import format_phone_for_display
from app.models import BusinessCardRecord

# image_provider(filename) -> raw image bytes, or None if the image is missing.
ImageProvider = Callable[[str], "bytes | None"]


def build_workbook_bytes(
    records: list[BusinessCardRecord],
    image_provider: ImageProvider | None = None,
) -> bytes:
    """Build the contacts workbook fully in memory and return the .xlsx bytes.

    Card images are fetched via ``image_provider`` (from GridFS) rather than read
    off disk, so nothing touches the local filesystem.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "contacts"
    ws.append([EXCEL_HEADERS.get(column, column) for column in EXCEL_COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")

    for record in records:
        values = record.model_dump() if hasattr(record, "model_dump") else record.dict()
        values["business"] = record.business or record.company
        values["date"] = f"{record.date} {record.time}".strip()
        values["email1"] = record.email1 or record.email
        values["contact1"] = format_phone_for_display(
            record.contact1 or record.mobile_number or record.phone_primary,
            record.country_code,
        )
        values["contact2"] = format_phone_for_display(record.contact2 or record.phone_number, record.country_code)
        values["contact3"] = format_phone_for_display(record.contact3 or record.fax_number, record.country_code)
        values["low_confidence_fields"] = ", ".join(record.low_confidence_fields)
        values["reviewed_by_user"] = "Yes" if record.reviewed_by_user else "No"
        values["card"] = ""
        ws.append([values.get(column) for column in EXCEL_COLUMNS])
        row_index = ws.max_row
        ws.row_dimensions[row_index].height = 92
        if record.front_image_filename and image_provider is not None:
            _add_card_image(ws, image_provider(record.front_image_filename), row_index, "card")
        fill = None
        if record.confidence_score == "Low":
            fill = PatternFill("solid", fgColor="FEE2E2")
        elif record.duplicate_flag != "No":
            fill = PatternFill("solid", fgColor="FEF9C3")
        if fill:
            for cell in ws[row_index]:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, column in enumerate(EXCEL_COLUMNS, start=1):
        width = 38 if column == "card" else min(max(len(EXCEL_HEADERS.get(column, column)) + 2, 14), 42)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def export_records(records: list[BusinessCardRecord], output_path) -> "Path":
    """Backward-compatible wrapper: write the workbook to a path on disk.

    The app itself streams the workbook from memory (build_workbook_bytes); this
    remains only for legacy/offline scripts. Images are not embedded here since
    they now live in GridFS, not on disk.
    """
    from pathlib import Path

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(build_workbook_bytes(records, image_provider=None))
    return output


# The sheet only ever displays this at a 140x84pt thumbnail (see image.width/
# height below), so the embedded pixel data only needs to comfortably exceed
# that on a high-DPI zoom — anything larger just bloats the .xlsx for no
# visible benefit. 320px covers a ~2.3x retina-ish zoom of the display box.
_CARD_THUMBNAIL_MAX_EDGE = 320
_CARD_THUMBNAIL_QUALITY = 75


def _add_card_image(ws, image_bytes: "bytes | None", row_index: int, column: str) -> None:
    if not image_bytes or column not in EXCEL_COLUMNS:
        return
    try:
        # Re-encode as a small JPEG rather than a lossless PNG: the source is
        # already a compressed JPEG (see compress_for_storage), so decoding it
        # and writing it back out as PNG would inflate a ~150-250KB file back
        # up to 1-1.5MB for a thumbnail nobody views above 140x84pt.
        with PillowImage.open(BytesIO(image_bytes)) as source:
            rgb = source.convert("RGB")
            width, height = rgb.size
            longest = max(width, height)
            if longest > _CARD_THUMBNAIL_MAX_EDGE:
                scale = _CARD_THUMBNAIL_MAX_EDGE / longest
                rgb = rgb.resize((max(1, int(width * scale)), max(1, int(height * scale))))
            buffer = BytesIO()
            rgb.save(buffer, "JPEG", quality=_CARD_THUMBNAIL_QUALITY, optimize=True)
        buffer.seek(0)
        image = ExcelImage(buffer)
    except Exception:  # noqa: BLE001 — a bad image should not fail the whole export
        return
    image.width = 140
    image.height = 84
    col_idx = EXCEL_COLUMNS.index(column) + 1
    ws.add_image(image, f"{get_column_letter(col_idx)}{row_index}")
